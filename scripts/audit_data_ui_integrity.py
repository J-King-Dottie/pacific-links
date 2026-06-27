#!/usr/bin/env python3
"""
Audit dashboard data integrity against the public CSVs and Excel export.

This checks the data layer that feeds the UI:
- one relationship direction per metric
- latest-mode rows use each pair's latest available year
- grouped detail rows sum back to parent rows
- toggled metric pairs are not accidentally identical
- Excel tabs match the public CSV aggregates
"""

from __future__ import annotations

import csv
import math
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PUBLIC = ROOT / "dashboard" / "public" / "data"
XLSX = PUBLIC / "pacific_links_data.xlsx"
YEAR_MIN = 2010
YEAR_MAX = 2024


@dataclass(frozen=True)
class MetricDef:
    metric: str
    tab: str
    csv_name: str
    value_col: str
    pct_col: str | None
    grouped: bool = False
    excel_id_cols: tuple[str, ...] = ("pacific_code", "pacific_name", "counterpart_code", "counterpart_name")


METRICS = [
    MetricDef("aid", "Aid", "aid_by_donor_year.csv", "value_usd", "pct_gdp"),
    MetricDef("aid_committed", "Aid committed", "aid_committed_by_donor_year.csv", "value_usd", "pct_gdp"),
    MetricDef("trade", "Imports", "imports_by_supplier_year.csv", "value_usd", "pct_gdp", grouped=True),
    MetricDef("exports", "Exports", "exports_by_destination_year.csv", "value_usd", "pct_gdp", grouped=True),
    MetricDef("debt", "Debt", "debt_by_creditor_year.csv", "value_usd", "pct_gdp"),
    MetricDef(
        "security",
        "Security assistance",
        "security_assistance_by_provider_year.csv",
        "value_usd",
        "pct_gdp",
        grouped=True,
        excel_id_cols=("pacific_code", "pacific_name", "counterpart_code", "counterpart_name", "sector_code", "sector_name"),
    ),
    MetricDef(
        "security_arms",
        "Security arms",
        "security_arms_by_supplier_year.csv",
        "value_tiv",
        None,
        grouped=True,
        excel_id_cols=("pacific_code", "pacific_name", "counterpart_code", "counterpart_name", "weapon_designation", "weapon_description", "delivery_years", "status"),
    ),
    MetricDef("remittances", "Remittances", "remittances_by_source_year.csv", "value_usd", "pct_gdp"),
    MetricDef("migration", "Migration", "migrants_abroad_by_destination_year.csv", "value_people", "pct_population"),
    MetricDef("students", "Students", "students_by_destination_year.csv", "value_people", "pct_population"),
    MetricDef("fdi", "FDI", "fdi_positions_by_investor_year.csv", "value_usd", "pct_gdp"),
    MetricDef("portfolio", "Portfolio", "portfolio_positions_by_holder_year.csv", "value_usd", "pct_gdp"),
]

EXPECTED_SHEETS = [
    "About",
    "Aid",
    "Aid committed",
    "Imports",
    "Exports",
    "Debt",
    "Security assistance",
    "Security arms",
    "Remittances",
    "Migration",
    "Students",
    "FDI",
    "Portfolio",
]

TOGGLES = [
    ("Aid spent vs promised", "aid", "aid_committed", "AU"),
    ("Imports vs exports", "trade", "exports", "AU"),
    ("Security assistance vs arms", "security", "security_arms", "AU"),
    ("FDI vs portfolio", "fdi", "portfolio", "AU"),
]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def fnum(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        n = float(value)
    except (TypeError, ValueError):
        return None
    return n if math.isfinite(n) else None


def dashboard_rows(metric: MetricDef) -> list[dict[str, object]]:
    rows = read_csv(PUBLIC / metric.csv_name)
    grouped: dict[tuple[str, str, str, str, int], dict[str, object]] = {}

    for r in rows:
        pac = r.get("pacific_code", "")
        pac_name = r.get("pacific_name", "")
        cp = r.get("counterpart_code", "")
        cp_name = r.get("counterpart_name", "")
        year = int(r.get("year") or 0)
        value = fnum(r.get(metric.value_col))
        pct = fnum(r.get(metric.pct_col)) if metric.pct_col else None
        if not pac or not cp or not value or value <= 0 or not (YEAR_MIN <= year <= YEAR_MAX):
            continue

        key = (pac, pac_name, cp, cp_name, year)
        if key not in grouped:
            grouped[key] = {
                "pacific_code": pac,
                "pacific_name": pac_name,
                "counterpart_code": cp,
                "counterpart_name": cp_name,
                "year": year,
                "value": 0.0,
                "pct": 0.0 if metric.pct_col else None,
                "detail_value": 0.0,
            }
        item = grouped[key]
        item["value"] = float(item["value"]) + value
        item["detail_value"] = float(item["detail_value"]) + value
        if metric.pct_col:
            item["pct"] = float(item["pct"] or 0) + (pct or 0)

    return list(grouped.values())


def latest_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    latest_year: dict[tuple[str, str], int] = {}
    for r in rows:
        key = (str(r["pacific_code"]), str(r["counterpart_code"]))
        latest_year[key] = max(latest_year.get(key, 0), int(r["year"]))
    return [r for r in rows if int(r["year"]) == latest_year[(str(r["pacific_code"]), str(r["counterpart_code"]))]]


def sum_value(rows: list[dict[str, object]]) -> float:
    return sum(float(r["value"]) for r in rows)


def fmt_money(value: float) -> str:
    if value >= 1e9:
        return f"${value / 1e9:.2f}B"
    if value >= 1e6:
        return f"${value / 1e6:.2f}M"
    if value >= 1e3:
        return f"${value / 1e3:.2f}K"
    return f"${value:.2f}"


def excel_expected(metric: MetricDef) -> dict[tuple[object, ...], dict[int, tuple[float | None, float | None]]]:
    expected: dict[tuple[object, ...], dict[int, tuple[float | None, float | None]]] = defaultdict(dict)
    rows = read_csv(PUBLIC / metric.csv_name)
    accum: dict[tuple[str, str, str, str, int], list[float | None]] = {}

    for r in rows:
        year = int(r.get("year") or 0)
        if not (YEAR_MIN <= year <= YEAR_MAX):
            continue
        key = tuple(r.get(col, "") for col in metric.excel_id_cols) + (year,)
        value = fnum(r.get(metric.value_col))
        pct = fnum(r.get(metric.pct_col)) if metric.pct_col else None
        if key not in accum:
            accum[key] = [value, pct]
        else:
            accum[key][0] = (accum[key][0] or 0) + value if value is not None else accum[key][0]
            if metric.pct_col:
                accum[key][1] = (accum[key][1] or 0) + pct if pct is not None else accum[key][1]

    for key5, values in accum.items():
        key = key5[:-1]
        year = key5[-1]
        expected[key][year] = (values[0], values[1])
    return expected


def audit_excel(metric: MetricDef, wb) -> list[str]:
    issues: list[str] = []
    expected = excel_expected(metric)
    if metric.tab not in wb.sheetnames:
        return [f"missing Excel tab: {metric.tab}"]

    ws = wb[metric.tab]
    headers = [c.value for c in ws[1]]
    id_width = len(metric.excel_id_cols)
    years = [int(h) for h in headers[id_width:] if isinstance(h, int) or (isinstance(h, str) and h.isdigit())]
    pct_start = id_width + len(years)
    sheet_keys = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        key = tuple(row[:id_width])
        sheet_keys.add(key)
        if key not in expected:
            issues.append(f"{metric.tab}: unexpected Excel row {key}")
            continue
        for i, year in enumerate(years):
            got = fnum(row[id_width + i])
            exp = expected[key].get(year, (None, None))[0]
            if (got or 0) == 0 and exp is None:
                continue
            if abs((got or 0) - (exp or 0)) > 0.05:
                issues.append(f"{metric.tab}: value mismatch {key} {year}: {got} != {exp}")
                break
        if metric.pct_col:
            for i, year in enumerate(years):
                got = fnum(row[pct_start + i])
                exp = expected[key].get(year, (None, None))[1]
                if (got or 0) == 0 and exp is None:
                    continue
                if abs((got or 0) - (exp or 0)) > 0.0002:
                    issues.append(f"{metric.tab}: pct mismatch {key} {year}: {got} != {exp}")
                    break

    missing = set(expected) - sheet_keys
    if missing:
        issues.append(f"{metric.tab}: {len(missing)} expected rows missing from Excel")
    return issues


def main() -> int:
    all_issues: list[str] = []
    summaries: list[str] = []
    latest_by_metric: dict[str, list[dict[str, object]]] = {}
    wb = load_workbook(XLSX, read_only=True, data_only=True)

    if wb.sheetnames != EXPECTED_SHEETS:
        all_issues.append(f"Excel sheet order mismatch: {wb.sheetnames}")

    for metric in METRICS:
        rows = dashboard_rows(metric)
        latest = latest_rows(rows)
        latest_by_metric[metric.metric] = latest
        if not rows:
            all_issues.append(f"{metric.metric}: no dashboard rows")
            continue

        if metric.grouped:
            for r in rows:
                if abs(float(r["value"]) - float(r["detail_value"])) > 0.05:
                    all_issues.append(f"{metric.metric}: detail sum mismatch for {r}")
                    break

        pac_count = len({r["pacific_code"] for r in latest})
        cp_count = len({r["counterpart_code"] for r in latest})
        year_min = min(int(r["year"]) for r in latest)
        year_max = max(int(r["year"]) for r in latest)
        total = sum_value(latest)
        summaries.append(
            f"- {metric.tab}: {len(rows):,} dashboard rows; latest view {len(latest):,} relationships; "
            f"{pac_count} Pacific countries; {cp_count} partners; latest years {year_min}-{year_max}; total {fmt_money(total)}"
        )

        all_issues.extend(audit_excel(metric, wb))

    for label, left, right, partner in TOGGLES:
        left_total = sum_value([r for r in latest_by_metric[left] if r["counterpart_code"] == partner])
        right_total = sum_value([r for r in latest_by_metric[right] if r["counterpart_code"] == partner])
        if left_total == right_total and left_total > 0:
            all_issues.append(f"{label}: {partner} totals are identical ({left_total})")
        summaries.append(f"- Toggle check {label} for {partner}: {fmt_money(left_total)} vs {fmt_money(right_total)}")

    for partner in ["AU", "NZ", "CN", "US"]:
        available = []
        for metric in METRICS:
            total = sum_value([r for r in latest_by_metric[metric.metric] if r["counterpart_code"] == partner])
            if total > 0:
                available.append(f"{metric.metric} {fmt_money(total)}")
        summaries.append(f"- Outside-partner latest footprint {partner}: " + ("; ".join(available) if available else "no rows"))

    print("# Data-to-UI Integrity Audit")
    print()
    print("## Summary")
    print("\n".join(summaries))
    print()
    print("## Result")
    if all_issues:
        for issue in all_issues:
            print(f"- FAIL: {issue}")
        return 1
    print("- PASS: public CSVs, latest-mode aggregates, grouped details, toggles, and Excel tabs are internally consistent.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
