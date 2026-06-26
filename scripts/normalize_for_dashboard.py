"""
Normalize processed CSVs for dashboard consumption.

Unified output schema for all metrics:
  pacific_code, pacific_name, counterpart_code, counterpart_name,
  year, value_usd, pct_gdp

Exceptions:
  - imports:   adds hs1_code, hs1_name
  - migration: value_people instead of value_usd; pct_population instead of pct_gdp

This script:
  1. Reads each data/processed CSV, cleans and renames columns
  2. Writes harmonized CSVs back to data/processed/ (canonical)
  3. Copies harmonized CSVs to dashboard/public/data/
  4. Generates dashboard/public/data/pacific_links_data.xlsx (one tab per metric,
     wide format: pacific + counterpart on left, year columns across the top)
"""

import csv
import json
import shutil
from collections import defaultdict
from datetime import datetime, date, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR    = Path(__file__).parent.parent / "data" / "processed"
RAW_DIR     = Path(__file__).parent.parent / "data" / "raw"
DASH_DIR    = Path(__file__).parent.parent / "dashboard" / "public" / "data"
EXCEL_PATH  = DASH_DIR / "pacific_links_data.xlsx"

EXCLUDED_PACIFIC = {"NC"}

# Stated scope of the project (dashboard + landing page + download).
# Sources carry ragged out-of-scope tails (aid back to 1983 and stray future
# "spent" years; migration back to 1990). Clip every output to this window so
# the download matches the claimed coverage. Numbers inside the window are
# untouched and still trace exactly to source.
YEAR_MIN = 2010
YEAR_MAX = 2024

REMITTANCE_SKIP_CODES = {"", "WORLD"}
REMITTANCE_SKIP_NAME_SUBSTRINGS = ["total", "unidentified", "other south"]

ISO3_TO_ISO2 = {
    "ABW": "AW", "ALB": "AL", "ARG": "AR", "ASM": "AS", "AUS": "AU",
    "AUT": "AT", "AZE": "AZ", "BEL": "BE", "BEN": "BJ", "BFA": "BF",
    "BGD": "BD", "BGR": "BG", "BIH": "BA", "BLR": "BY", "BLZ": "BZ",
    "BMU": "BM", "BOL": "BO", "BRA": "BR", "BRB": "BB", "BWA": "BW",
    "CAN": "CA", "CHE": "CH", "CHL": "CL", "CHN": "CN", "COL": "CO",
    "CRI": "CR", "CUB": "CU", "CYM": "KY", "CYP": "CY", "CZE": "CZ",
    "DEU": "DE", "DNK": "DK", "DOM": "DO", "DZA": "DZ", "ECU": "EC",
    "EGY": "EG", "ESP": "ES", "EST": "EE", "FIN": "FI", "FJI": "FJ",
    "FRA": "FR", "FSM": "FM", "GBR": "GB", "GIN": "GN", "GNB": "GW",
    "GRC": "GR", "GTM": "GT", "GUM": "GU", "HKG": "HK", "HRV": "HR",
    "HUN": "HU", "IDN": "ID", "IND": "IN", "IRL": "IE", "ISL": "IS",
    "ISR": "IL", "ITA": "IT", "JOR": "JO", "JPN": "JP", "KAZ": "KZ",
    "KEN": "KE", "KIR": "KI", "KOR": "KR", "KOS": "XK", "KWT": "KW",
    "LBN": "LB", "LKA": "LK", "LTU": "LT", "LUX": "LU", "MAC": "MO",
    "MEX": "MX", "MHL": "MH", "MKD": "MK", "MLI": "ML", "MLT": "MT",
    "MNP": "MP", "MOZ": "MZ", "MUS": "MU", "NGA": "NG",
    "NIC": "NI", "NLD": "NL", "NOR": "NO", "NZL": "NZ", "PAK": "PK",
    "PAN": "PA", "PER": "PE", "PHL": "PH", "PLW": "PW", "PNG": "PG",
    "POL": "PL", "PRI": "PR", "PRT": "PT", "PYF": "PF", "ROU": "RO",
    "RUS": "RU", "SEN": "SN", "SGP": "SG", "SLB": "SB", "SLV": "SV",
    "SRB": "RS", "SVK": "SK", "SVN": "SI", "SWE": "SE", "SWZ": "SZ",
    "SYC": "SC", "TGO": "TG", "THA": "TH", "TON": "TO", "TUR": "TR",
    "TUV": "TV", "TWN": "TW", "UKR": "UA", "URY": "UY", "USA": "US",
    "VEN": "VE", "VGB": "VG", "VIR": "VI", "VNM": "VN", "VUT": "VU",
    "WSM": "WS", "ZAF": "ZA",
}

M49_TO_ISO2 = {
    "16": "AS",  "36": "AU",  "68": "BO",  "90": "SB",  "100": "BG",
    "124": "CA", "152": "CL", "170": "CO", "184": "CK", "188": "CR",
    "196": "CY", "208": "DK", "214": "DO", "233": "EE", "242": "FJ",
    "246": "FI", "258": "PF", "296": "KI", "300": "GR", "316": "GU",
    "324": "GN", "348": "HU", "352": "IS", "356": "IN", "428": "LV",
    "440": "LT", "442": "LU", "466": "ML", "484": "MX", "520": "NR",
    "548": "VU", "554": "NZ", "570": "NU", "578": "NO",
    "580": "MP", "583": "FM", "584": "MH", "585": "PW", "598": "PG",
    "703": "SK", "705": "SI", "710": "ZA", "772": "TK", "776": "TO",
    "798": "TV", "876": "WF", "882": "WS",
}


def load_gdp_pop():
    lookup = {}
    for r in csv.DictReader(open(RAW_DIR / "worldbank_gdp_pop.csv", encoding="utf-8")):
        iso2 = r["iso2"]
        year = int(r["year"])
        gdp  = float(r["gdp_usd"])    if r.get("gdp_usd")    else None
        pop  = float(r["population"]) if r.get("population") else None
        lookup.setdefault(iso2, {})[year] = {"gdp_usd": gdp, "population": pop}
    return lookup


def best_gdp(gdp_pop, iso2, year):
    country = gdp_pop.get(iso2, {})
    if year in country and country[year]["gdp_usd"]:
        return country[year]["gdp_usd"]
    for delta in range(1, 4):
        for y in [year - delta, year + delta]:
            if y in country and country[y]["gdp_usd"]:
                return country[y]["gdp_usd"]
    return None


def best_pop(gdp_pop, iso2, year):
    country = gdp_pop.get(iso2, {})
    if year in country and country[year]["population"]:
        return country[year]["population"]
    for delta in range(1, 4):
        for y in [year - delta, year + delta]:
            if y in country and country[y]["population"]:
                return country[y]["population"]
    return None


def write_csv(path, fieldnames, rows):
    if not rows:
        raise RuntimeError(f"Refusing to write 0 rows to {path}")
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


# ---------------------------------------------------------------------------
# Normalizers — each returns list of dicts in the unified schema
# ---------------------------------------------------------------------------

def normalize_aid(gdp_pop):
    rows = list(csv.DictReader(open(DATA_DIR / "aid_by_donor_year.csv", encoding="utf-8")))
    out = []
    for r in rows:
        pac  = r.get("pacific_code") or r.get("recipient_code", "")
        cpty = r.get("counterpart_code") or r.get("donor_code", "")
        try:
            val = float(r.get("aid_spent_usd") or r.get("value_usd") or 0)
        except ValueError:
            continue
        if val <= 0 or pac in EXCLUDED_PACIFIC:
            continue
        year = int(r["year"])
        gdp  = best_gdp(gdp_pop, pac, year)
        out.append({
            "pacific_code":     pac,
            "pacific_name":     r.get("pacific_name") or r.get("recipient_name", ""),
            "counterpart_code": cpty,
            "counterpart_name": r.get("counterpart_name") or r.get("donor_name", ""),
            "year":             year,
            "value_usd":        round(val, 2),
            "pct_gdp":          round(val / gdp * 100, 4) if gdp else None,
        })
    print(f"  Aid:          {len(out):>6} rows")
    return out


def normalize_imports(gdp_pop):
    baci_path     = RAW_DIR / "baci_imports.csv"
    comtrade_path = RAW_DIR / "comtrade_mirror_imports.csv"
    pdh_path      = DATA_DIR / "imports_by_supplier_year.csv"
    src = baci_path if baci_path.exists() else (comtrade_path if comtrade_path.exists() else pdh_path)
    print(f"    imports source: {src.name}")
    rows = list(csv.DictReader(open(src, encoding="utf-8")))
    out = []
    for r in rows:
        pac   = r.get("pacific_code") or r.get("reporter_code", "")
        cpty  = r.get("counterpart_code") or r.get("supplier_code", "")
        if cpty == "RESIDUAL":
            continue
        try:
            val = float(r.get("value_usd") or r.get("import_value_usd") or 0)
        except (ValueError, KeyError):
            continue
        if val <= 0 or pac in EXCLUDED_PACIFIC:
            continue
        year = int(r["year"])
        gdp  = best_gdp(gdp_pop, pac, year)
        out.append({
            "pacific_code":     pac,
            "pacific_name":     r.get("pacific_name") or r.get("reporter_name", ""),
            "counterpart_code": cpty,
            "counterpart_name": r.get("counterpart_name") or r.get("supplier_name", ""),
            "year":             year,
            "hs1_code":         r.get("hs1_code", ""),
            "hs1_name":         r.get("hs1_name", ""),
            "value_usd":        round(val, 2),
            "pct_gdp":          round(val / gdp * 100, 4) if gdp else None,
        })
    print(f"  Imports:      {len(out):>6} rows")
    return out


def normalize_remittances(gdp_pop):
    rows = list(csv.DictReader(open(DATA_DIR / "remittances_by_source_year.csv", encoding="utf-8")))
    out = []
    dropped = 0
    for r in rows:
        pac = r.get("pacific_code") or r.get("recipient_code", "")
        # source_code may be ISO3, or already resolved to ISO2 counterpart_code
        raw_code = r.get("source_code", r.get("counterpart_code", "")).strip()
        name = (r.get("source_name") or r.get("counterpart_name", "")).strip().lower()
        if raw_code in REMITTANCE_SKIP_CODES or any(s in name for s in REMITTANCE_SKIP_NAME_SUBSTRINGS):
            dropped += 1
            continue
        try:
            val = float(r.get("value_usd") or r.get("remittance_estimate_usd") or 0)
        except (ValueError, KeyError):
            continue
        if val <= 0 or pac in EXCLUDED_PACIFIC:
            continue
        iso2 = r.get("counterpart_code") or r.get("source_iso2", "").strip() or ISO3_TO_ISO2.get(raw_code, "")
        year = int(r["year"])
        gdp  = best_gdp(gdp_pop, pac, year)
        out.append({
            "pacific_code":     pac,
            "pacific_name":     r.get("pacific_name") or r.get("recipient_name", ""),
            "counterpart_code": iso2,
            "counterpart_name": r.get("counterpart_name") or r.get("source_name", ""),
            "year":             year,
            "value_usd":        round(val, 2),
            "pct_gdp":          round(val / gdp * 100, 4) if gdp else None,
        })
    print(f"  Remittances:  {len(out):>6} rows (dropped {dropped} aggregates)")
    return out


def normalize_migration(gdp_pop):
    rows = list(csv.DictReader(open(DATA_DIR / "migrants_abroad_by_destination_year.csv", encoding="utf-8")))
    valid = []
    for r in rows:
        pac = r.get("pacific_code") or r.get("origin_code", "")
        raw = r.get("destination_code", "").strip()
        iso2 = r.get("counterpart_code") or r.get("destination_iso2", "").strip() or M49_TO_ISO2.get(raw, "")
        try:
            val = float(r.get("value_people") or r.get("value") or r.get("migrant_stock") or 0)
        except (ValueError, KeyError):
            continue
        if val <= 0 or pac in EXCLUDED_PACIFIC:
            continue
        valid.append({
            "pacific_code":     pac,
            "pacific_name":     r.get("pacific_name") or r.get("origin_name", ""),
            "counterpart_code": iso2,
            "counterpart_name": r.get("counterpart_name") or r.get("destination_name", ""),
            "year":             int(r["year"]),
            "value_people":     round(val, 2),
        })

    # pct_population = migrants to this destination / total origin population
    out = []
    for r in valid:
        pop = best_pop(gdp_pop, r["pacific_code"], r["year"])
        out.append({**r, "pct_population": round(r["value_people"] / pop * 100, 4) if pop else None})

    print(f"  Migration:    {len(out):>6} rows")
    return out


def normalize_debt(gdp_pop):
    rows = list(csv.DictReader(open(DATA_DIR / "debt_by_creditor_year.csv", encoding="utf-8")))
    out = []
    for r in rows:
        pac  = r.get("pacific_code") or r.get("recipient_code", "")
        cpty = r.get("counterpart_code") or r.get("creditor_code", "")
        try:
            val = float(r.get("value_usd") or 0)
        except ValueError:
            continue
        if val <= 0 or pac in EXCLUDED_PACIFIC:
            continue
        year = int(r["year"])
        gdp  = best_gdp(gdp_pop, pac, year)
        out.append({
            "pacific_code":     pac,
            "pacific_name":     r.get("pacific_name") or r.get("recipient_name", ""),
            "counterpart_code": cpty,
            "counterpart_name": r.get("counterpart_name") or r.get("creditor_name", ""),
            "year":             year,
            "value_usd":        round(val, 2),
            "pct_gdp":          round(val / gdp * 100, 4) if gdp else None,
        })
    print(f"  Debt:         {len(out):>6} rows")
    return out


# ---------------------------------------------------------------------------
# Source metadata — mirrors METRIC_INFO in SidePanel.jsx
# ---------------------------------------------------------------------------

METRIC_META = [
    {
        "metric":    "Aid",
        "tab":       "Aid",
        "source":    "Lowy Institute Pacific Aid Map via Pacific Data Hub",
        "sourceUrl": "https://pacificdata.org/data/dataset/pacific-aid-and-development-finance-data-from-the-lowy-institute-df-pam",
        "sourceLinks": [],
        "coverage":  "14 Pacific island countries, 2010-2024 where aid records are available.",
        "value_col": "value_usd",
        "value_label": "USD (spent/disbursed aid)",
        "pct_col":   "pct_gdp",
        "pct_label": "% of recipient GDP",
        "notes":     "Spent/disbursed aid only — not commitments. Aggregate donor categories removed.",
    },
    {
        "metric":    "Imports",
        "tab":       "Imports",
        "source":    "CEPII BACI Reconciled Bilateral Trade Data",
        "sourceUrl": "https://www.cepii.fr/CEPII/en/bdd_modele/bdd_modele_item.asp?id=37",
        "sourceLinks": [],
        "coverage":  "14 Pacific island countries, 2010-2024 where import records are available.",
        "value_col": "value_usd",
        "value_label": "USD (import value)",
        "pct_col":   "pct_gdp",
        "pct_label": "% of importer GDP",
        "notes":     "Grouped by supplier country. Small flows tied to ship registration / bunkering fuel removed (HS 8901, HS 271000 for MH, HS 89 for MH).",
    },
    {
        "metric":    "Remittances",
        "tab":       "Remittances",
        "source":    "World Bank and KNOMAD Bilateral Remittance Matrices",
        "sourceUrl": "https://web.archive.org/web/20231119232039/https://www.knomad.org/data/remittances",
        "sourceLinks": [
            ("2021 matrix", "https://thedocs.worldbank.org/en/doc/cf8eee7ff5029398f75e897b342e7320-0050122023/related/WB-KNOMAD.xlsx"),
            ("2018 matrix", "https://thedocs.worldbank.org/en/doc/904591573826885707-0090022019/original/Bilateralremittancematrix2018Oct2019.xlsx"),
            ("2017 matrix (archive)", "https://web.archive.org/web/20190124090151id_/http://pubdocs.worldbank.org/en/705611533661084197/bilateralremittancematrix2017-Apr2018.xlsx"),
            ("2010 matrix (archive)", "https://web.archive.org/web/20160531235700id_/http://pubdocs.worldbank.org:80/pubdocs/publicdoc/2015/9/895701443117529385/Bilateral-Remittance-Matrix-2010.xlsx"),
        ],
        "coverage":  "14 Pacific island countries. Benchmark years only: 2010, 2017, 2018, 2021.",
        "value_col": "value_usd",
        "value_label": "USD (modelled bilateral estimate)",
        "pct_col":   "pct_gdp",
        "pct_label": "% of recipient GDP",
        "notes":     "Modelled bilateral estimates — not observed flows. No interpolation between benchmark years. Non-country aggregate rows removed.",
    },
    {
        "metric":    "Migration",
        "tab":       "Migration",
        "source":    "UN International Migrant Stock 2024",
        "sourceUrl": "https://www.un.org/development/desa/pd/content/international-migrant-stock",
        "sourceLinks": [],
        "coverage":  "14 Pacific island countries. Benchmark years only: 1990, 1995, 2000, 2005, 2010, 2015, 2020, 2024.",
        "value_col": "value_people",
        "value_label": "People (migrant stock — people born in Pacific country living in destination)",
        "pct_col":   "pct_population",
        "pct_label": "% of origin country population (can exceed 100% for small islands)",
        "notes":     "Migrant stock, not annual flows. Regional and aggregate destination rows removed.",
    },
    {
        "metric":    "Debt",
        "tab":       "Debt",
        "source":    "World Bank International Debt Statistics (IDS)",
        "sourceUrl": "https://databank.worldbank.org/source/international-debt-statistics",
        "sourceLinks": [],
        "coverage":  "Fiji, FSM, PNG, Samoa, Solomon Islands, Tonga, Vanuatu — IDS reporting countries only, 2010-2024.",
        "value_col": "value_usd",
        "value_label": "USD (external debt stock — public and publicly guaranteed)",
        "pct_col":   "pct_gdp",
        "pct_label": "% of debtor GDP",
        "notes":     "Debt stock (outstanding obligations), not new borrowing. IDS creditor names retained including multilateral institutions.",
    },
]


# ---------------------------------------------------------------------------
# Source vintage / freshness
# ---------------------------------------------------------------------------
# Each metric carries a "release" label (the source edition we pulled) and a way
# to date it. Dates come from each build script's metadata.json (retrieved_at)
# where available, falling back to the mtime of the relevant raw/processed file.

BACI_VERSION = "HS92_V202601"  # keep in sync with scripts/fetch_baci_trade.py

SOURCE_RELEASES = {
    "Aid":         {"release": "Lowy Pacific Aid Map (live SDMX)",          "meta": "aid_by_donor_year.metadata.json"},
    "Imports":     {"release": f"CEPII BACI {BACI_VERSION.replace('_', ' ')}", "file": RAW_DIR / "baci" / f"BACI_{BACI_VERSION}.zip"},
    "Remittances": {"release": "World Bank/KNOMAD bilateral matrices",      "meta": "remittances_by_source_year.metadata.json"},
    "Migration":   {"release": "UN International Migrant Stock 2024",       "meta": "migrants_abroad_by_destination_year.metadata.json"},
    "Debt":        {"release": "World Bank IDS (live API)",                 "file": DATA_DIR / "debt_by_creditor_year.csv"},
}


def _date_from_meta(name):
    path = DATA_DIR / name
    try:
        stamp = json.loads(path.read_text(encoding="utf-8")).get("retrieved_at", "")
        return stamp[:10] or None
    except Exception:
        return None


def _date_from_mtime(path):
    try:
        return datetime.fromtimestamp(Path(path).stat().st_mtime, tz=timezone.utc).date().isoformat()
    except Exception:
        return None


def collect_freshness():
    """Return (last_refreshed, [ {metric, release, retrieved} ]) for manifest + Excel."""
    sources = []
    for m in METRIC_META:
        cfg = SOURCE_RELEASES.get(m["metric"], {})
        retrieved = _date_from_meta(cfg["meta"]) if cfg.get("meta") else _date_from_mtime(cfg.get("file"))
        sources.append({
            "metric":    m["metric"],
            "release":   cfg.get("release", m["source"]),
            "retrieved": retrieved or "unknown",
        })
    last_refreshed = date.today().isoformat()
    return last_refreshed, sources


# ---------------------------------------------------------------------------
# Excel export — one tab per metric, wide format (years as columns)
# ---------------------------------------------------------------------------

HEADER_FILL  = PatternFill("solid", fgColor="2A6B72")
HEADER_FONT  = Font(bold=True, color="FFFAF0", size=10)
ALT_FILL     = PatternFill("solid", fgColor="F0EBE0")
THIN         = Side(style="thin", color="C4B090")
BORDER       = Border(bottom=Side(style="thin", color="C4B090"))

TAB_COLOR = {
    "Aid":          "2A6B72",
    "Imports":      "8A5C10",
    "Remittances":  "4A7A50",
    "Migration":    "5A4080",
    "Debt":         "A04030",
}


def _header_style(ws, row, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center" if col > 4 else "left", vertical="center")


DENOMINATOR_LINKS = [
    ("GDP — World Bank (most countries)",    "https://data.worldbank.org/indicator/NY.GDP.MKTP.CD"),
    ("GDP — UN SNAAMA (Cook Islands)",       "https://unstats.un.org/unsd/snaama/"),
    ("GDP — Niue Statistics Office (Niue)",  "https://niue.prism.spc.int/"),
    ("Population — World Bank",              "https://data.worldbank.org/indicator/SP.POP.TOTL"),
]


def build_about_sheet(wb, last_refreshed, freshness):
    ws = wb.create_sheet(title="About")
    ws.sheet_properties.tabColor = "2A6B72"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 85

    TITLE_FONT   = Font(bold=True, size=13, color="2A6B72")
    SECTION_FONT = Font(bold=True, size=9,  color="8A5C10")
    SOURCE_FONT  = Font(bold=True, size=9,  color="2C1F0A")
    LINK_FONT    = Font(size=9,   color="2A6B72", underline="single")
    DIM_FONT     = Font(size=8,   color="9A8060")
    BODY_FONT    = Font(size=9,   color="2C1F0A")

    def add(r, a="", b="", a_font=None, b_font=None, b_link=None, h=None):
        ca = ws.cell(row=r, column=1, value=a)
        cb = ws.cell(row=r, column=2, value=b)
        if a_font: ca.font = a_font
        if b_link:
            cb.value = b
            cb.hyperlink = b_link
            cb.font = LINK_FONT
        elif b_font:
            cb.font = b_font
        if h: ws.row_dimensions[r].height = h
        return r + 1

    fresh_by_metric = {f["metric"]: f for f in freshness}

    r = 1
    r = add(r, "Pacific Links — Data Sources", h=20, a_font=TITLE_FONT)
    r = add(r, "Compiled and harmonized by Dottie AI Studio", a_font=DIM_FONT)
    r = add(r, f"Data last refreshed: {last_refreshed}", a_font=Font(bold=True, size=9, color="2A6B72"))
    r = add(r)

    r = add(r, "Data sources", a_font=SECTION_FONT)
    for m in METRIC_META:
        f = fresh_by_metric.get(m["metric"], {})
        r = add(r, m["metric"], m["source"], a_font=SOURCE_FONT, b_font=BODY_FONT)
        r = add(r, "", m["sourceUrl"], b_link=m["sourceUrl"])
        for label, url in m.get("sourceLinks", []):
            r = add(r, f"  {label}", url, a_font=DIM_FONT, b_link=url)
        r = add(r, "  edition / pulled", f"{f.get('release','')}  ({f.get('retrieved','unknown')})", a_font=DIM_FONT, b_font=DIM_FONT)
        r = add(r)

    r = add(r, "Denominator sources", a_font=SECTION_FONT)
    for label, url in DENOMINATOR_LINKS:
        r = add(r, label, url, a_font=Font(size=9, color="2C1F0A", bold=True), b_link=url)


def build_excel_tab(wb, tab_name, rows, value_col, pct_col, id_cols, id_labels):
    """
    id_cols:   list of row dict keys that form the unique row identifier
    id_labels: human-readable column headers for id_cols
    value_col: key in row dict for the primary value
    pct_col:   key in row dict for the % value (None if not applicable)
    """
    # Collect all years present
    years = sorted({int(r["year"]) for r in rows})

    # Build lookup: (id tuple) -> {year: [value_sum, pct_sum]}
    # Sum across duplicate (key, year) rows. This matters for Imports, where each
    # supplier-year has multiple HS1 category rows that must be totalled, not
    # overwritten. For the other metrics there is one row per key-year, so the
    # sum is a no-op.
    data = defaultdict(dict)
    id_sets = {}
    for r in rows:
        key = tuple(r[c] for c in id_cols)
        id_sets[key] = key
        yr = int(r["year"])
        v = r.get(value_col)
        p = r.get(pct_col)
        v = float(v) if v not in (None, "") else None
        p = float(p) if p not in (None, "") else None
        if yr not in data[key]:
            data[key][yr] = [v, p]
        else:
            cur = data[key][yr]
            cur[0] = (cur[0] or 0) + v if v is not None else cur[0]
            cur[1] = (cur[1] or 0) + p if p is not None else cur[1]

    sorted_keys = sorted(id_sets.keys())

    ws = wb.create_sheet(title=tab_name)
    if tab_name in TAB_COLOR:
        ws.sheet_properties.tabColor = TAB_COLOR[tab_name]

    # Header row
    headers = id_labels + [str(y) for y in years]
    if pct_col:
        headers += [f"{y} % GDP" if pct_col == "pct_gdp" else f"{y} % pop" for y in years]

    for col_i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center" if col_i > len(id_labels) else "left",
            vertical="center",
        )

    ws.row_dimensions[1].height = 18

    # Data rows
    for row_i, key in enumerate(sorted_keys, 2):
        fill = ALT_FILL if row_i % 2 == 0 else None
        # ID columns
        for col_i, val in enumerate(key, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            if fill:
                cell.fill = fill
        # Value columns
        for yr_i, yr in enumerate(years):
            col_i = len(id_labels) + 1 + yr_i
            val, pct = data[key].get(yr, (None, None))
            cell = ws.cell(row=row_i, column=col_i, value=val if val is not None else None)
            cell.number_format = '#,##0'
            if fill:
                cell.fill = fill
        # Pct columns
        if pct_col:
            for yr_i, yr in enumerate(years):
                col_i = len(id_labels) + 1 + len(years) + yr_i
                val, pct = data[key].get(yr, (None, None))
                cell = ws.cell(row=row_i, column=col_i, value=round(pct, 4) if pct is not None else None)
                cell.number_format = '0.00"%"'
                if fill:
                    cell.fill = fill

    # Column widths
    id_widths = [8, 22, 12, 28]  # approximate for code/name pairs
    for col_i, w in enumerate(id_widths[:len(id_labels)], 1):
        ws.column_dimensions[get_column_letter(col_i)].width = w
    for col_i in range(len(id_labels) + 1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 11

    ws.freeze_panes = ws.cell(row=2, column=len(id_labels) + 1)
    print(f"  Excel tab '{tab_name}': {len(sorted_keys)} rows × {len(years)} years")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

DATASETS = [
    # (filename_stem, fieldnames)
    ("aid_by_donor_year",               ["pacific_code", "pacific_name", "counterpart_code", "counterpart_name", "year", "value_usd", "pct_gdp"]),
    ("imports_by_supplier_year",        ["pacific_code", "pacific_name", "counterpart_code", "counterpart_name", "year", "hs1_code", "hs1_name", "value_usd", "pct_gdp"]),
    ("remittances_by_source_year",      ["pacific_code", "pacific_name", "counterpart_code", "counterpart_name", "year", "value_usd", "pct_gdp"]),
    ("migrants_abroad_by_destination_year", ["pacific_code", "pacific_name", "counterpart_code", "counterpart_name", "year", "value_people", "pct_population"]),
    ("debt_by_creditor_year",           ["pacific_code", "pacific_name", "counterpart_code", "counterpart_name", "year", "value_usd", "pct_gdp"]),
]


if __name__ == "__main__":
    print("Normalizing and harmonizing CSVs...")
    gdp_pop = load_gdp_pop()
    print(f"  GDP/pop loaded for {len(gdp_pop)} countries\n")

    aid         = normalize_aid(gdp_pop)
    imports     = normalize_imports(gdp_pop)
    remittances = normalize_remittances(gdp_pop)
    migration   = normalize_migration(gdp_pop)
    debt        = normalize_debt(gdp_pop)

    # Clip all outputs to the stated 2010-2024 scope.
    def clip(rows):
        return [r for r in rows if YEAR_MIN <= int(r["year"]) <= YEAR_MAX]
    before = {"aid": len(aid), "imports": len(imports), "remittances": len(remittances),
              "migration": len(migration), "debt": len(debt)}
    aid, imports, remittances, migration, debt = map(clip, (aid, imports, remittances, migration, debt))
    print(f"\nClipped to {YEAR_MIN}-{YEAR_MAX}: "
          f"aid {before['aid']}->{len(aid)}, migration {before['migration']}->{len(migration)} "
          f"(others unchanged)")

    datasets = [
        ("aid_by_donor_year",                   aid,         ["pacific_code", "pacific_name", "counterpart_code", "counterpart_name", "year", "value_usd", "pct_gdp"]),
        ("imports_by_supplier_year",             imports,     ["pacific_code", "pacific_name", "counterpart_code", "counterpart_name", "year", "hs1_code", "hs1_name", "value_usd", "pct_gdp"]),
        ("remittances_by_source_year",           remittances, ["pacific_code", "pacific_name", "counterpart_code", "counterpart_name", "year", "value_usd", "pct_gdp"]),
        ("migrants_abroad_by_destination_year",  migration,   ["pacific_code", "pacific_name", "counterpart_code", "counterpart_name", "year", "value_people", "pct_population"]),
        ("debt_by_creditor_year",                debt,        ["pacific_code", "pacific_name", "counterpart_code", "counterpart_name", "year", "value_usd", "pct_gdp"]),
    ]

    print("\nWriting harmonized CSVs...")
    DASH_DIR.mkdir(parents=True, exist_ok=True)
    for stem, rows, fieldnames in datasets:
        processed_path = DATA_DIR / f"{stem}.csv"
        dashboard_path = DASH_DIR / f"{stem}.csv"
        write_csv(processed_path, fieldnames, rows)
        shutil.copy2(processed_path, dashboard_path)
        print(f"  {stem}.csv -> processed/ + dashboard/public/data/")

    last_refreshed, freshness = collect_freshness()
    meta_path = DASH_DIR / "data_meta.json"
    meta_path.write_text(json.dumps({
        "last_refreshed": last_refreshed,
        "year_min": YEAR_MIN,
        "year_max": YEAR_MAX,
        "sources": freshness,
    }, indent=2) + "\n", encoding="utf-8")
    print(f"\nData last refreshed: {last_refreshed}")
    for f in freshness:
        print(f"  {f['metric']:12} {f['release']}  ({f['retrieved']})")

    print("\nBuilding Excel download...")
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_about_sheet(wb, last_refreshed, freshness)

    id_cols   = ["pacific_code", "pacific_name", "counterpart_code", "counterpart_name"]
    id_labels = ["Pacific ISO2", "Pacific Country", "Counterpart Code", "Counterpart Name"]

    build_excel_tab(wb, "Aid",         aid,         "value_usd",    "pct_gdp",       id_cols, id_labels)
    build_excel_tab(wb, "Imports",     imports,     "value_usd",    "pct_gdp",       id_cols, id_labels)
    build_excel_tab(wb, "Remittances", remittances, "value_usd",    "pct_gdp",       id_cols, id_labels)
    build_excel_tab(wb, "Migration",   migration,   "value_people", "pct_population", id_cols, id_labels)
    build_excel_tab(wb, "Debt",        debt,        "value_usd",    "pct_gdp",       id_cols, id_labels)

    wb.save(EXCEL_PATH)
    print(f"\n  Saved: {EXCEL_PATH}")
    print("\nDone.")
