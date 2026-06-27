#!/usr/bin/env python3
"""
Audit provenance and reproducibility for published dashboard data.

This does not verify that source publishers are correct. It checks that every
published metric has a traceable source story across:
- the normalizer metadata used by the app and Excel export
- dashboard/public/data/data_meta.json
- public CSV files
- the Excel download and About tab
- repository docs
"""

from __future__ import annotations

import csv
import json
import re
import runpy
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PUBLIC = ROOT / "dashboard" / "public" / "data"
DATA_DIR = ROOT / "data" / "processed"
NORMALIZER = ROOT / "scripts" / "normalize_for_dashboard.py"
XLSX = PUBLIC / "pacific_links_data.xlsx"
README = ROOT / "README.md"
PIPELINE = ROOT / "DATA_PIPELINE.md"


@dataclass(frozen=True)
class PublishedMetric:
    metric: str
    tab: str
    csv_name: str
    value_col: str
    source_keyword: str


PUBLISHED = [
    PublishedMetric("Aid", "Aid", "aid_by_donor_year.csv", "value_usd", "Lowy"),
    PublishedMetric("Aid committed", "Aid committed", "aid_committed_by_donor_year.csv", "value_usd", "Lowy"),
    PublishedMetric("Imports", "Imports", "imports_by_supplier_year.csv", "value_usd", "BACI"),
    PublishedMetric("Exports", "Exports", "exports_by_destination_year.csv", "value_usd", "BACI"),
    PublishedMetric("Debt", "Debt", "debt_by_creditor_year.csv", "value_usd", "IDS"),
    PublishedMetric("Security assistance", "Security assistance", "security_assistance_by_provider_year.csv", "value_usd", "OECD"),
    PublishedMetric("Security arms", "Security arms", "security_arms_by_supplier_year.csv", "value_tiv", "SIPRI"),
    PublishedMetric("Remittances", "Remittances", "remittances_by_source_year.csv", "value_usd", "KNOMAD"),
    PublishedMetric("Migration", "Migration", "migrants_abroad_by_destination_year.csv", "value_people", "Migrant Stock"),
    PublishedMetric("Students", "Students", "students_by_destination_year.csv", "value_people", "UNESCO"),
    PublishedMetric("FDI", "FDI", "fdi_positions_by_investor_year.csv", "value_usd", "Direct Investment"),
    PublishedMetric("Portfolio", "Portfolio", "portfolio_positions_by_holder_year.csv", "value_usd", "Portfolio Investment"),
]

DOC_SOURCE_TERMS = [
    "Lowy",
    "CEPII BACI",
    "World Bank IDS",
    "OECD CRS",
    "SIPRI",
    "KNOMAD",
    "UN International Migrant Stock",
    "UNESCO UIS",
    "IMF Direct Investment",
    "IMF Portfolio Investment",
]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def load_normalizer_context() -> dict[str, object]:
    return runpy.run_path(str(NORMALIZER))


def date_like(value: str) -> bool:
    return bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}", value or ""))


def excel_about_rows() -> list[tuple[object, object]]:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    rows = [(a, b) for a, b, *_ in ws_rows(wb["About"])]
    wb.close()
    return rows


def ws_rows(ws):
    for row in ws.iter_rows(values_only=True):
        yield row


def main() -> int:
    issues: list[str] = []
    summary: list[str] = []

    ctx = load_normalizer_context()
    metric_meta = ctx.get("METRIC_META", [])
    source_releases = ctx.get("SOURCE_RELEASES", {})
    meta_by_metric = {m.get("metric"): m for m in metric_meta}

    expected_metrics = [m.metric for m in PUBLISHED]
    if set(meta_by_metric) != set(expected_metrics):
        issues.append(f"METRIC_META set mismatch: {sorted(meta_by_metric)}")

    source_manifest = json.loads((PUBLIC / "data_meta.json").read_text(encoding="utf-8"))
    manifest_sources = source_manifest.get("sources", [])
    manifest_by_metric = {s.get("metric"): s for s in manifest_sources}
    if set(manifest_by_metric) != set(expected_metrics):
        issues.append(f"data_meta.json source set mismatch: {sorted(manifest_by_metric)}")
    if source_manifest.get("year_min") != 2010 or source_manifest.get("year_max") != 2024:
        issues.append("data_meta.json year window is not 2010-2024")

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    sheetnames = wb.sheetnames
    about_text = "\n".join(
        " ".join(str(cell) for cell in row if cell is not None)
        for row in wb["About"].iter_rows(values_only=True)
    )
    wb.close()

    docs = {
        "README.md": README.read_text(encoding="utf-8"),
        "DATA_PIPELINE.md": PIPELINE.read_text(encoding="utf-8"),
    }

    for metric in PUBLISHED:
        m = meta_by_metric.get(metric.metric)
        if not m:
            issues.append(f"{metric.metric}: missing METRIC_META entry")
            continue

        for field in ("source", "sourceUrl", "coverage", "value_col", "value_label", "notes"):
            if not str(m.get(field, "")).strip():
                issues.append(f"{metric.metric}: METRIC_META missing {field}")
        if not str(m.get("sourceUrl", "")).startswith("http"):
            issues.append(f"{metric.metric}: sourceUrl is not an HTTP URL")
        if m.get("tab") != metric.tab:
            issues.append(f"{metric.metric}: tab mismatch {m.get('tab')} != {metric.tab}")
        if m.get("value_col") != metric.value_col:
            issues.append(f"{metric.metric}: value column mismatch {m.get('value_col')} != {metric.value_col}")
        if len(str(m.get("notes", ""))) < 80:
            issues.append(f"{metric.metric}: notes look too thin for methodology/provenance")

        release_cfg = source_releases.get(metric.metric)
        manifest = manifest_by_metric.get(metric.metric)
        if not release_cfg:
            issues.append(f"{metric.metric}: missing SOURCE_RELEASES entry")
        if not manifest:
            issues.append(f"{metric.metric}: missing data_meta.json source entry")
        elif release_cfg:
            if manifest.get("release") != release_cfg.get("release"):
                issues.append(f"{metric.metric}: release mismatch between SOURCE_RELEASES and data_meta.json")
            if not date_like(manifest.get("retrieved", "")):
                issues.append(f"{metric.metric}: retrieved date is missing or not YYYY-MM-DD")

        csv_path = PUBLIC / metric.csv_name
        if not csv_path.exists():
            issues.append(f"{metric.metric}: missing public CSV {metric.csv_name}")
        else:
            rows = read_csv(csv_path)
            if not rows:
                issues.append(f"{metric.metric}: public CSV has no rows")
            elif metric.value_col not in rows[0]:
                issues.append(f"{metric.metric}: public CSV missing value column {metric.value_col}")
            summary.append(f"- {metric.metric}: public CSV {len(rows):,} rows; source {m.get('source')}")

        processed_path = DATA_DIR / metric.csv_name
        if not processed_path.exists():
            issues.append(f"{metric.metric}: missing processed CSV {metric.csv_name}")

        if release_cfg and release_cfg.get("meta"):
            processed_meta = DATA_DIR / str(release_cfg["meta"])
            if not processed_meta.exists():
                issues.append(f"{metric.metric}: missing processed metadata {processed_meta.name}")
            else:
                raw = json.loads(processed_meta.read_text(encoding="utf-8"))
                if not (raw.get("source") or raw.get("source_dataset") or raw.get("dataset")):
                    issues.append(f"{metric.metric}: processed metadata lacks source/dataset field")
                if not str(raw.get("retrieved_at", "")).strip():
                    issues.append(f"{metric.metric}: processed metadata lacks retrieved_at")
        elif release_cfg and release_cfg.get("file"):
            if not Path(release_cfg["file"]).exists():
                issues.append(f"{metric.metric}: SOURCE_RELEASES fallback file is missing")

        if metric.tab not in sheetnames:
            issues.append(f"{metric.metric}: Excel tab {metric.tab} missing")
        for needle in (metric.metric, str(m.get("source", "")), str(m.get("sourceUrl", ""))):
            if needle and needle not in about_text:
                issues.append(f"{metric.metric}: Excel About tab missing {needle}")
        if manifest and manifest.get("release") and manifest["release"] not in about_text:
            issues.append(f"{metric.metric}: Excel About tab missing release {manifest['release']}")
        if metric.source_keyword not in about_text:
            issues.append(f"{metric.metric}: Excel About tab missing expected keyword {metric.source_keyword}")

    for doc_name, text in docs.items():
        for term in DOC_SOURCE_TERMS:
            if term not in text:
                issues.append(f"{doc_name}: missing source term {term}")

    print("# Provenance and Reproducibility Audit")
    print()
    print("## Summary")
    print("\n".join(summary))
    print()
    print("## Result")
    if issues:
        for issue in issues:
            print(f"- FAIL: {issue}")
        return 1
    print("- PASS: published metrics have aligned source metadata, public CSVs, Excel tabs, About-sheet citations, and repo documentation.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
