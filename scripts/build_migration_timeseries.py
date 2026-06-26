from __future__ import annotations

import csv
import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

RAW_PATH = Path("data/raw/un_migrant_stock_bilateral_2024.xlsx")
OUTPUT_DIR = Path("data/processed")
CSV_PATH = OUTPUT_DIR / "migrants_abroad_by_destination_year.csv"
METADATA_PATH = OUTPUT_DIR / "migrants_abroad_by_destination_year.metadata.json"
SOURCE_URL = (
    "https://www.un.org/development/desa/pd/sites/"
    "www.un.org.development.desa.pd/files/undesa_pd_2024_ims_stock_by_sex_destination_and_origin.xlsx"
)

PACIFIC_ORIGINS = {
    "American Samoa": "AS",
    "Cook Islands": "CK",
    "Fiji": "FJ",
    "French Polynesia": "PF",
    "Guam": "GU",
    "Kiribati": "KI",
    "Marshall Islands": "MH",
    "Micronesia (Fed. States of)": "FM",
    "Nauru": "NR",
    "Niue": "NU",
    "Northern Mariana Islands": "MP",
    "Palau": "PW",
    "Papua New Guinea": "PG",
    "Samoa": "WS",
    "Solomon Islands": "SB",
    "Tokelau": "TK",
    "Tonga": "TO",
    "Tuvalu": "TV",
    "Vanuatu": "VU",
    "Wallis and Futuna Islands": "WF",
}

YEAR_COLUMNS = {
    1990: 7,
    1995: 8,
    2000: 9,
    2005: 10,
    2010: 11,
    2015: 12,
    2020: 13,
    2024: 14,
}


def clean_name(value: object) -> str:
    return " ".join(str(value or "").replace("*", "").split()).strip()


def positive_number(value: object) -> float | None:
    if isinstance(value, (int, float)) and value > 0:
        return float(value)
    return None


def build_dataset() -> list[dict[str, object]]:
    wb = load_workbook(RAW_PATH, read_only=True, data_only=True)
    ws = wb["Table 1"]

    rows: list[dict[str, object]] = []
    for values in ws.iter_rows(min_row=12, values_only=True):
        destination_name = clean_name(values[1])
        destination_code = values[4]
        origin_name = clean_name(values[5])
        origin_code_num = values[6]

        if origin_name not in PACIFIC_ORIGINS:
            continue
        if not isinstance(destination_code, int) or destination_code >= 900:
            continue
        if not isinstance(origin_code_num, int) or origin_code_num >= 900:
            continue

        for year, col in YEAR_COLUMNS.items():
            value = positive_number(values[col])
            if value is None:
                continue
            rows.append(
                {
                    "origin_code": PACIFIC_ORIGINS[origin_name],
                    "origin_name": origin_name,
                    "destination_code": str(destination_code),
                    "destination_name": destination_name,
                    "year": year,
                    "migrant_stock": round(value, 2),
                    "destination_share_pct": "",
                    "source_workbook": RAW_PATH.name,
                    "source_url": SOURCE_URL,
                }
            )

    totals = defaultdict(float)
    for row in rows:
        totals[(str(row["origin_code"]), int(row["year"]))] += float(row["migrant_stock"])
    for row in rows:
        total = totals[(str(row["origin_code"]), int(row["year"]))]
        row["destination_share_pct"] = round(float(row["migrant_stock"]) / total * 100, 6)

    rows.sort(key=lambda row: (row["origin_code"], row["year"], -float(row["migrant_stock"]), row["destination_name"]))
    return rows


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    rows = build_dataset()
    write_csv(CSV_PATH, rows)

    coverage = {}
    for code in sorted({str(row["origin_code"]) for row in rows}):
        subset = [row for row in rows if row["origin_code"] == code]
        coverage[code] = {
            "years": sorted({int(row["year"]) for row in subset}),
            "destination_country_count": len({str(row["destination_name"]) for row in subset}),
        }

    metadata = {
        "title": "Pacific-origin migrant stock abroad by destination country",
        "source": "United Nations International Migrant Stock 2024: Destination and origin",
        "retrieved_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "row_count": len(rows),
        "origin_count": len(coverage),
        "years": sorted({int(row["year"]) for row in rows}),
        "coverage": coverage,
        "unavailable_pacific_members": sorted(set(PACIFIC_ORIGINS.values()) - set(coverage)),
        "notes": [
            "Values represent estimated migrant stock, not migration flows.",
            "The source workbook provides benchmark years only: 1990, 1995, 2000, 2005, 2010, 2015, 2020, 2024.",
            "This extract keeps Pacific countries and territories as origins and country or area destinations with positive stock values.",
            "Destination shares are calculated separately within each origin and benchmark year.",
        ],
        "source_url": SOURCE_URL,
    }
    METADATA_PATH.write_text(json.dumps(metadata, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(metadata, indent=2))


if __name__ == "__main__":
    main()
