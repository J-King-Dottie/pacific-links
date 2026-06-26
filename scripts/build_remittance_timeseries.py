from __future__ import annotations

import csv
import argparse
import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

RAW_DIR = Path("data/raw")
OUTPUT_DIR = Path("data/processed")
CSV_PATH = OUTPUT_DIR / "remittances_by_source_year.csv"
METADATA_PATH = OUTPUT_DIR / "remittances_by_source_year.metadata.json"
STAGING_DIR = OUTPUT_DIR / "remittance_benchmarks"

PACIFIC = {
    "American Samoa": "AS", "Cook Islands": "CK", "Fiji": "FJ",
    "French Polynesia": "PF", "Micronesia, Fed. Sts.": "FM",
    "Micronesia, Federated States of": "FM", "Guam": "GU",
    "Kiribati": "KI", "Marshall Islands": "MH", "Nauru": "NR",
    "Niue": "NU", "Northern Mariana Islands": "MP",
    "Papua New Guinea": "PG", "Palau": "PW", "Samoa": "WS",
    "Solomon Islands": "SB", "Tokelau": "TK", "Tonga": "TO",
    "Tuvalu": "TV", "Vanuatu": "VU", "Wallis and Futuna": "WF",
    "Pitcairn": "PN", "Pitcairn Islands": "PN",
}

FILES = {
    2010: RAW_DIR / "world_bank_bilateral_remittances_2010.xlsx",
    2017: RAW_DIR / "world_bank_bilateral_remittances_2017.xlsx",
    2018: RAW_DIR / "world_bank_bilateral_remittances_2018.xlsx",
    2021: RAW_DIR / "world_bank_knomad.xlsx",
}

SOURCE_URLS = {
    2010: "https://web.archive.org/web/20160531235700id_/http://pubdocs.worldbank.org:80/pubdocs/publicdoc/2015/9/895701443117529385/Bilateral-Remittance-Matrix-2010.xlsx",
    2017: "https://web.archive.org/web/20190124090151id_/http://pubdocs.worldbank.org:80/en/705611533661084197/bilateralremittancematrix2017-Apr2018.xlsx",
    2018: "https://thedocs.worldbank.org/en/doc/904591573826885707-0090022019/original/Bilateralremittancematrix2018Oct2019.xlsx",
    2021: "https://thedocs.worldbank.org/en/doc/cf8eee7ff5029398f75e897b342e7320-0050122023/related/WB-KNOMAD.xlsx",
}


def clean_name(value: object) -> str:
    return " ".join(str(value or "").replace("’", "'").split()).strip()


def positive_number(value: object) -> float | None:
    if isinstance(value, (int, float)) and value > 0:
        return float(value)
    return None


def country_code_lookup() -> dict[str, str]:
    """Build a name-to-ISO3 lookup from the 2018 workbook's coded axes."""
    wb = load_workbook(FILES[2018], read_only=True, data_only=True)
    ws = wb.active
    lookup = {}
    header_names = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    header_codes = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    for col in range(2, len(header_names)):
        name, code = clean_name(header_names[col]), clean_name(header_codes[col])
        if name and len(code) == 3:
            lookup[name] = code
    for values in ws.iter_rows(min_row=4, values_only=True):
        name, code = clean_name(values[0]), clean_name(values[1])
        if name and len(code) == 3:
            lookup[name] = code
    return lookup


def read_matrix(year: int, source_codes: dict[str, str]) -> list[dict[str, object]]:
    wb = load_workbook(FILES[year], read_only=True, data_only=True)
    ws = wb.active
    if year == 2018:
        header_row, data_row, first_recipient_col, source_name_col, source_code_col = 2, 4, 3, 1, 2
    else:
        header_row, data_row, first_recipient_col, source_name_col, source_code_col = 2, 3, 2, 1, None

    recipients = []
    header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    for col in range(first_recipient_col - 1, len(header_values)):
        name = clean_name(header_values[col])
        if name in PACIFIC:
            recipients.append((col, name, PACIFIC[name]))

    rows = []
    for values in ws.iter_rows(min_row=data_row, values_only=True):
        source_name = clean_name(values[source_name_col - 1])
        source_code = clean_name(values[source_code_col - 1]) if source_code_col else source_codes.get(source_name, "")
        if not source_name:
            continue
        for col, recipient_name, recipient_code in recipients:
            value_millions = positive_number(values[col])
            if value_millions is None:
                continue
            rows.append({
                "recipient_code": recipient_code,
                "recipient_name": recipient_name,
                "source_code": source_code,
                "source_name": source_name,
                "year": year,
                "remittance_estimate_usd": round(value_millions * 1_000_000, 2),
                "source_share_pct": "",
                "source_workbook": FILES[year].name,
                "source_url": SOURCE_URLS[year],
            })
    return rows


def read_2021() -> list[dict[str, object]]:
    wb = load_workbook(FILES[2021], read_only=True, data_only=True)
    ws = wb["Data"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    index = {str(value): pos for pos, value in enumerate(headers)}
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if values[index["Indicator ID"]] != "WB.KNOMAD.BRE":
            continue
        recipient_name = clean_name(values[index["Partner"]])
        if recipient_name not in PACIFIC:
            continue
        value_millions = positive_number(values[index["2021"]])
        if value_millions is None:
            continue
        rows.append({
            "recipient_code": PACIFIC[recipient_name],
            "recipient_name": recipient_name,
            "source_code": clean_name(values[index["Economy ISO3"]]),
            "source_name": clean_name(values[index["Economy Name"]]),
            "year": 2021,
            "remittance_estimate_usd": round(value_millions * 1_000_000, 2),
            "source_share_pct": "",
            "source_workbook": FILES[2021].name,
            "source_url": SOURCE_URLS[2021],
        })
    return rows


def write_rows(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--year", type=int, choices=FILES)
    args = parser.parse_args()

    if args.year:
        rows = read_2021() if args.year == 2021 else read_matrix(
            args.year, country_code_lookup() if args.year in (2010, 2017) else {}
        )
        write_rows(STAGING_DIR / f"remittances_{args.year}.csv", rows)
        print(f"Saved {len(rows)} rows for {args.year}")
        return

    rows = []
    for year in FILES:
        path = STAGING_DIR / f"remittances_{year}.csv"
        if not path.exists():
            raise FileNotFoundError(f"Run --year {year} first: {path}")
        with path.open(encoding="utf-8", newline="") as handle:
            rows.extend(csv.DictReader(handle))

    keys = [(row["recipient_code"], row["source_name"], row["year"]) for row in rows]
    if len(keys) != len(set(keys)):
        raise RuntimeError(f"Found {len(keys) - len(set(keys))} duplicate recipient-source-year rows")

    totals = defaultdict(float)
    for row in rows:
        totals[(str(row["recipient_code"]), int(row["year"]))] += float(row["remittance_estimate_usd"])
    for row in rows:
        total = totals[(str(row["recipient_code"]), int(row["year"]))]
        row["source_share_pct"] = round(float(row["remittance_estimate_usd"]) / total * 100, 6)
    rows.sort(key=lambda row: (row["recipient_code"], row["year"], row["source_name"]))

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    write_rows(CSV_PATH, rows)

    coverage = {}
    for code in sorted({str(row["recipient_code"]) for row in rows}):
        subset = [row for row in rows if row["recipient_code"] == code]
        coverage[code] = {
            "years": sorted({int(row["year"]) for row in subset}),
            "source_country_count": len({str(row["source_name"]) for row in subset}),
        }
    metadata = {
        "title": "Modelled bilateral remittances received by Pacific destination and source country",
        "source": "World Bank/KNOMAD Bilateral Remittance Matrices",
        "retrieved_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "row_count": len(rows),
        "recipient_count": len(coverage),
        "years": sorted({int(row["year"]) for row in rows}),
        "coverage": coverage,
        "unavailable_pacific_members": sorted(set(PACIFIC.values()) - set(coverage)),
        "notes": [
            "All bilateral values are analytical model estimates, not observed transaction totals.",
            "Values are converted from current USD millions to current USD.",
            "Only published benchmark years are included; missing years are not interpolated.",
            "Source shares are calculated separately within each recipient and benchmark year.",
            "The 2010 and 2017 live World Bank URLs were retired; preserved official workbooks were retrieved from the Internet Archive.",
        ],
        "source_urls": SOURCE_URLS,
    }
    METADATA_PATH.write_text(json.dumps(metadata, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(metadata, indent=2))


if __name__ == "__main__":
    main()
